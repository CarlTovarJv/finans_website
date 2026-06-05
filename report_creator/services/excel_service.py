from io import BytesIO
import pandas as pd

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter


def generate_excel(data, insights=None):

    buffer = BytesIO()

    df = pd.DataFrame(data)

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        # =========================
        # REPORT SHEET
        # =========================
        df.to_excel(
            writer,
            index=False,
            sheet_name="Financial Report"
        )

        workbook = writer.book
        worksheet = writer.sheets["Financial Report"]

        header_fill = PatternFill(
            "solid",
            fgColor="010221"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
            size=12
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Encabezados
        for cell in worksheet[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center"
            )
            cell.border = thin_border

        # Bordes y alineación
        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:

                cell.border = thin_border

                cell.alignment = Alignment(
                    horizontal="center"
                )

        # Ajustar ancho automático
        for column in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + 4

            worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width

        # =========================
        # SUMMARY SHEET
        # =========================
        if insights:

            summary = workbook.create_sheet(
                "Financial Summary"
            )

            summary["A1"] = "FINANCIAL SUMMARY"

            summary["A1"].font = Font(
                bold=True,
                size=16,
                color="FFFFFF"
            )

            summary["A1"].fill = PatternFill(
                "solid",
                fgColor="010221"
            )

            summary.merge_cells("A1:B1")

            metrics = [
                ["Total Income", insights["income"]],
                ["Total Sales", insights["sales"]],
                ["Total Revenue", insights["revenue"]],
                ["Total Expenses", insights["expenses"]],
                ["Net Profit", insights["profit"]],
                ["Profit Margin (%)", insights["profit_margin"]],
                ["Status", insights["message"]]
            ]

            start_row = 3

            for i, item in enumerate(metrics):

                row = start_row + i

                summary.cell(
                    row=row,
                    column=1,
                    value=item[0]
                )

                summary.cell(
                    row=row,
                    column=2,
                    value=item[1]
                )

            # Estilo tabla resumen
            for row in range(
                start_row,
                start_row + len(metrics)
            ):

                summary.cell(
                    row=row,
                    column=1
                ).font = Font(
                    bold=True
                )

                summary.cell(
                    row=row,
                    column=1
                ).fill = PatternFill(
                    "solid",
                    fgColor="E8ECF8"
                )

                summary.cell(
                    row=row,
                    column=1
                ).border = thin_border

                summary.cell(
                    row=row,
                    column=2
                ).border = thin_border

            # Color de Profit
            profit_cell = summary["B7"]

            if insights["profit"] >= 0:

                profit_cell.fill = PatternFill(
                    "solid",
                    fgColor="C6EFCE"
                )

            else:

                profit_cell.fill = PatternFill(
                    "solid",
                    fgColor="FFC7CE"
                )

            summary.column_dimensions["A"].width = 30
            summary.column_dimensions["B"].width = 25

    buffer.seek(0)

    return buffer