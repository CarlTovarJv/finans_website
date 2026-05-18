# services/excel_service.py

import pandas as pd

from openpyxl import load_workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter


def generate_excel(data, filename):

    rows = []

    total = 0

    for item in data:

        amount = float(item.get("amount") or 0)

        unit_price = float(item.get("unit_price") or 0)

        quantity = int(item.get("quantity") or 0)

        total += amount

        rows.append({

            "Type": item.get("type", ""),

            "Product": item.get("product", ""),

            "Quantity": quantity,

            "Unit Price": unit_price,

            "Amount": amount,

            "Date": item.get("date", "")

        })

    df = pd.DataFrame(rows)

    # =========================
    # TOTAL ROW
    # =========================

    df.loc[len(df)] = {

        "Type": "",

        "Product": "TOTAL",

        "Quantity": "",

        "Unit Price": "",

        "Amount": total,

        "Date": ""

    }

    df.to_excel(

        filename,

        index=False,

        sheet_name="Financial Report"

    )

    wb = load_workbook(filename)

    ws = wb["Financial Report"]

    # =========================
    # COLORS
    # =========================

    dark_blue = PatternFill(

        start_color="0F172A",

        end_color="0F172A",

        fill_type="solid"

    )

    cyan = PatternFill(

        start_color="06B6D4",

        end_color="06B6D4",

        fill_type="solid"

    )

    total_fill = PatternFill(

        start_color="DBEAFE",

        end_color="DBEAFE",

        fill_type="solid"

    )

    white_font = Font(

        color="FFFFFF",

        bold=True,

        size=12

    )

    bold_font = Font(

        bold=True

    )

    # =========================
    # BLACK BORDERS
    # =========================

    thin = Side(

        border_style="thin",

        color="000000"

    )

    border = Border(

        left=thin,

        right=thin,

        top=thin,

        bottom=thin

    )

    # =========================
    # TITLE
    # =========================

    ws.insert_rows(1)

    ws.merge_cells("A1:F1")

    title = ws["A1"]

    title.value = "KUALI AI FINANCIAL REPORT"

    title.font = Font(

        bold=True,

        size=20,

        color="FFFFFF"

    )

    title.fill = cyan

    title.alignment = Alignment(

        horizontal="center",

        vertical="center"

    )

    ws.row_dimensions[1].height = 35

    # =========================
    # HEADERS
    # =========================

    for cell in ws[2]:

        cell.fill = dark_blue

        cell.font = white_font

        cell.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )

        cell.border = border

    # =========================
    # BODY
    # =========================

    for row in ws.iter_rows(

        min_row=3,

        max_row=ws.max_row

    ):

        for cell in row:

            cell.border = border

            cell.alignment = Alignment(

                horizontal="center",

                vertical="center"

            )

    # =========================
    # MONEY FORMAT
    # =========================

    for row in ws.iter_rows(

        min_row=3,

        max_row=ws.max_row,

        min_col=4,

        max_col=5

    ):

        for cell in row:

            if isinstance(

                cell.value,

                (int, float)

            ):

                cell.number_format = '$#,##0.00'

    # =========================
    # TOTAL ROW STYLE
    # =========================

    for cell in ws[ws.max_row]:

        cell.fill = total_fill

        cell.font = bold_font

        cell.border = border

    # =========================
    # AUTO WIDTH
    # =========================

    for column_cells in ws.columns:

        max_length = 0

        column = column_cells[0].column

        column_letter = get_column_letter(column)

        for cell in column_cells:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(str(cell.value))

            except:
                pass

        adjusted_width = max_length + 6

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    wb.save(filename)

    return filename